from openpyxl import load_workbook
from datetime import datetime
import math

# ✅ Replace this path if your file is elsewhere
wb = load_workbook(r'C:\Users\PeniPendar\Desktop\Final Project\sagatave_eksamenam.xlsx', data_only=True)
ws = wb['Lapa_0']

# Read headers from row 3
header_row_index = 3
headers = {cell.value.strip(): idx for idx, cell in enumerate(ws[header_row_index], 1)}

# Initialize counters
count_ain_skaits = 0
count_high_2015 = 0
count_adulienas = 0
laserjet_sum = 0
laserjet_count = 0
korporativais_sum = 0

# Process data from row 4 onward
for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
    adrese = row[headers['Adrese'] - 1]
    skaits = row[headers['Skaits'] - 1]
    prioritate = row[headers['Prioritāte'] - 1]
    pieg_datums = row[headers['Piegādes datums'] - 1]
    pilseta = row[headers['Pilsēta'] - 1]
    produkts = row[headers['Produkts'] - 1]
    cena = row[headers['Cena'] - 1]
    klients = row[headers['Klients'] - 1]
    kopa = row[headers['Kopā'] - 1]

    # 1. Adrese starts with "Ain" and Skaits < 40
    if adrese and isinstance(skaits, (int, float)) and adrese.startswith("Ain") and skaits < 40:
        count_ain_skaits += 1

    # 2. Prioritāte is High and year is 2015
    if prioritate == 'High' and isinstance(pieg_datums, datetime) and pieg_datums.year == 2015:
        count_high_2015 += 1

    # 3. Adrese contains "Adulienas iela" and Pilsēta is Valmiera or Saulkrasti
    if adrese and 'Adulienas iela' in adrese and pilseta in ['Valmiera', 'Saulkrasti']:
        count_adulienas += 1

    # 4. Produkts contains "LaserJet", collect Cena
    if produkts and 'LaserJet' in produkts and isinstance(cena, (int, float)):
        laserjet_sum += cena
        laserjet_count += 1

    # 5. Klients is Korporatīvais and Skaits between 40–50
    if klients == 'Korporatīvais' and isinstance(skaits, (int, float)) and 40 <= skaits <= 50 and isinstance(kopa, (int, float)):
        korporativais_sum += kopa

# Final results
average_laserjet = math.floor(laserjet_sum / laserjet_count) if laserjet_count else 0
total_korporativais = math.floor(korporativais_sum)

# Output results
print("1. Entries where Adrese starts with 'Ain' and Skaits < 40:", count_ain_skaits)
print("2. Entries with Prioritāte = High and Piegādes datums in 2015:", count_high_2015)
print("3. Entries with Adulienas iela in Valmiera or Saulkrasti:", count_adulienas)
print("4. Average Cena for LaserJet products (rounded down):", average_laserjet)
print("5. Total Kopā for Korporatīvais clients (Skaits 40–50, rounded down):", total_korporativais)
